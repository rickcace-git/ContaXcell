"""Pruebas de la importación, la exportación y el guardado en disco.

La prueba que más vale de todas es la de ida y vuelta: importar un libro,
exportarlo y volver a importarlo tiene que dar exactamente lo mismo. Si eso se
cumple, ningún dato se queda por el camino.
"""

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

from contaxcell import calculos, excel  # noqa: E402
from contaxcell.almacen import Almacen  # noqa: E402
from contaxcell.modelo import (  # noqa: E402
    GASTO, INGRESO, INVERSION,
    Activo, AportacionGratis, Categoria, Libro, Movimiento, Valoracion,
)

# Los libros de ejemplo están en la carpeta de arriba, junto a la aplicación
# del móvil. Si alguien se lleva solo la carpeta `escritorio`, estas pruebas
# se saltan en vez de fallar.
LIBROS = [RAIZ.parent / "PlantillaContabilidad.xlsx",
          RAIZ.parent / "ContabilidadRicardo.xlsx"]
DISPONIBLES = [p for p in LIBROS if p.exists()]
PLANTILLA = RAIZ.parent / "PlantillaContabilidad.xlsx"


def retrato(libro: Libro) -> tuple:
    """Todo lo que define un libro, en algo comparable con ==."""
    cartera = calculos.cartera(libro)
    return (
        libro.ajustes.saldo_inicial,
        libro.ajustes.objetivo_inversion,
        [(c.nombre, c.tipo, c.presupuesto) for c in libro.categorias],
        sorted((m.fecha, m.descripcion, m.categoria, m.importe, m.activo)
               for m in libro.movimientos),
        [(a.nombre, a.aportacion_inicial, a.valor_mercado, a.ultima_valoracion)
         for a in libro.activos],
        sorted((g.fecha, g.activo, g.concepto, g.importe)
               for g in libro.aportaciones_gratis),
        sorted((v.fecha, v.valor_mercado) for v in libro.historico),
        calculos.saldo_banco(libro),
        cartera.total_aportado,
        cartera.generado,
    )


def libro_completo() -> Libro:
    libro = Libro.vacio()
    libro.ajustes.saldo_inicial = 1234.56
    libro.ajustes.objetivo_inversion = 150.0
    libro.categorias[2].presupuesto = 0.0      # un hueco a propósito
    libro.categorias[3].presupuesto = 480.0
    libro.categorias[5].presupuesto = 0.0
    libro.categorias[6].presupuesto = 90.0
    libro.activos = [
        Activo(nombre="Fondo indexado", aportacion_inicial=100.0, valor_mercado=265.0,
               ultima_valoracion="2026-08-31"),
        Activo(nombre="Cuenta remunerada", aportacion_inicial=0.0, valor_mercado=50.0,
               ultima_valoracion="2026-08-31"),
    ]
    libro.movimientos = [
        Movimiento(fecha="2026-08-01", descripcion="Nómina", categoria="Sueldo",
                   importe=1500.0),
        Movimiento(fecha="2026-08-03", descripcion="Alquiler",
                   categoria="Vivienda y Suministros", importe=600.0),
        Movimiento(fecha="2026-08-15", descripcion="Aportación", categoria="Inversión",
                   importe=120.0, activo="Fondo indexado"),
        Movimiento(fecha="2026-08-20", descripcion="Sin activo", categoria="Inversión",
                   importe=30.0),
    ]
    libro.aportaciones_gratis = [
        AportacionGratis(fecha="2026-08-31", activo="Fondo indexado",
                         concepto="Cashback 1%", importe=8.83),
    ]
    libro.historico = [Valoracion(fecha="2026-08-31", valor_mercado=315.0)]
    return libro


class PruebasIdaYVuelta(unittest.TestCase):
    def setUp(self):
        self.carpeta = Path(tempfile.mkdtemp(prefix="contaxcell-pruebas-"))

    def test_libro_construido_a_mano(self):
        original = libro_completo()
        destino = self.carpeta / "vuelta.xlsx"
        excel.exportar(destino, original, 2026)
        recuperado, _ = excel.importar(destino)
        self.assertEqual(retrato(recuperado), retrato(original))

    @unittest.skipUnless(DISPONIBLES, "no hay libros de ejemplo al lado")
    def test_libros_de_ejemplo(self):
        for ruta in DISPONIBLES:
            with self.subTest(libro=ruta.name):
                original, _ = excel.importar(ruta)
                destino = self.carpeta / f"vuelta-{ruta.stem}.xlsx"
                excel.exportar(destino, original, 2026)
                recuperado, _ = excel.importar(destino)
                self.assertEqual(retrato(recuperado), retrato(original))

    def test_conserva_los_huecos_del_presupuesto(self):
        # Una categoría sin tope entre dos que sí lo tienen no puede cortar la
        # lectura: las de abajo perderían el suyo.
        original = libro_completo()
        destino = self.carpeta / "presupuesto.xlsx"
        excel.exportar(destino, original, 2026)
        recuperado, _ = excel.importar(destino)
        self.assertEqual([c.presupuesto for c in recuperado.categorias],
                         [c.presupuesto for c in original.categorias])

    def test_conserva_el_tipo_de_cada_categoria(self):
        original = libro_completo()
        destino = self.carpeta / "tipos.xlsx"
        excel.exportar(destino, original, 2026)
        recuperado, _ = excel.importar(destino)
        tipos = {c.nombre: c.tipo for c in recuperado.categorias}
        self.assertEqual(tipos["Sueldo"], INGRESO)
        self.assertEqual(tipos["Inversión"], INVERSION)
        self.assertEqual(tipos["Transporte"], GASTO)


@unittest.skipUnless(DISPONIBLES, "no hay libros de ejemplo al lado")
class PruebasImportacionReal(unittest.TestCase):
    def test_la_plantilla_se_lee_entera(self):
        ruta = next(p for p in DISPONIBLES if p.name == "PlantillaContabilidad.xlsx")
        libro, avisos = excel.importar(ruta)
        self.assertEqual(libro.ajustes.saldo_inicial, 1000.0)
        self.assertEqual(len(libro.categorias), 9)
        self.assertEqual(libro.categoria("Inversión").tipo, INVERSION)
        self.assertEqual(libro.categoria("Vivienda y Suministros").presupuesto, 700.0)
        self.assertTrue(libro.movimientos)
        self.assertTrue(libro.activos)
        self.assertEqual(avisos, [])

    def test_el_balance_de_la_hoja_cuadra_con_el_calculado(self):
        # La columna G de la plantilla es una fórmula que la aplicación ignora
        # y recalcula. Tienen que dar lo mismo.
        for ruta in DISPONIBLES:
            with self.subTest(libro=ruta.name):
                libro, _ = excel.importar(ruta)
                filas = calculos.con_balance(libro)
                a_mano = libro.ajustes.saldo_inicial
                for fila in filas:
                    signo = 1 if fila.tipo == INGRESO else -1
                    a_mano = round(a_mano + signo * fila.importe, 2)
                self.assertAlmostEqual(filas[-1].balance, a_mano, places=2)


class PruebasImportacionDificil(unittest.TestCase):
    def setUp(self):
        self.carpeta = Path(tempfile.mkdtemp(prefix="contaxcell-pruebas-"))

    def test_archivo_sin_hoja_movimientos(self):
        from openpyxl import Workbook
        cuaderno = Workbook()
        cuaderno.active.title = "Hoja1"
        ruta = self.carpeta / "otro.xlsx"
        cuaderno.save(ruta)

        with self.assertRaises(excel.ErrorDeImportacion) as caso:
            excel.importar(ruta)
        self.assertIn("Movimientos", str(caso.exception))

    def test_archivo_que_no_es_excel(self):
        ruta = self.carpeta / "no-es-excel.xlsx"
        ruta.write_text("esto no es un libro", encoding="utf-8")
        with self.assertRaises(excel.ErrorDeImportacion):
            excel.importar(ruta)

    def test_categoria_que_no_esta_en_el_panel(self):
        from openpyxl import Workbook
        cuaderno = Workbook()
        hoja = cuaderno.active
        hoja.title = "Movimientos"
        hoja["A1"] = "Fecha"
        hoja["A2"] = "2026-05-04"
        hoja["C2"] = "Gimnasio"
        hoja["D2"] = "Deporte"
        hoja["F2"] = 30
        hoja["J6"] = "Sueldo"
        hoja["K6"] = "Ingreso"
        ruta = self.carpeta / "categoria-suelta.xlsx"
        cuaderno.save(ruta)

        libro, avisos = excel.importar(ruta)
        self.assertEqual(libro.categoria("Deporte").tipo, GASTO)
        self.assertTrue(any("Deporte" in a for a in avisos))

    def test_fechas_en_varios_formatos(self):
        self.assertEqual(excel._a_fecha("24/08/2026"), "2026-08-24")
        self.assertEqual(excel._a_fecha("4-8-26"), "2026-08-04")
        self.assertEqual(excel._a_fecha("2026-08-24"), "2026-08-24")
        self.assertEqual(excel._a_fecha("mañana"), "")
        self.assertEqual(excel._a_fecha(None), "")

    def test_importes_con_simbolos(self):
        self.assertEqual(excel._a_numero("1234,56"), 1234.56)
        self.assertEqual(excel._a_numero(" 40 € "), 40.0)
        self.assertEqual(excel._a_numero(None), 0.0)
        self.assertEqual(excel._a_numero("nada"), 0.0)


class PruebasExportacionSobreLaPlantilla(unittest.TestCase):
    """La exportación ya no fabrica una hoja: rellena la plantilla original.

    La prueba reina es la de identidad: importar la plantilla y exportarla
    tiene que devolver un archivo idéntico celda a celda, fórmulas incluidas.
    """

    def setUp(self):
        self.carpeta = Path(tempfile.mkdtemp(prefix="contaxcell-plantilla-"))

    @unittest.skipUnless(PLANTILLA.exists(), "no está la plantilla al lado")
    def test_exportar_la_plantilla_la_reproduce_identica(self):
        from openpyxl import load_workbook

        libro, _ = excel.importar(PLANTILLA)
        destino = self.carpeta / "identica.xlsx"
        excel.exportar(destino, libro, 2026)

        original = load_workbook(PLANTILLA)
        copia = load_workbook(destino)
        self.assertEqual(original.sheetnames, copia.sheetnames)
        for nombre in original.sheetnames:
            hoja_a, hoja_b = original[nombre], copia[nombre]
            self.assertEqual(
                sorted(str(r) for r in hoja_a.merged_cells.ranges),
                sorted(str(r) for r in hoja_b.merged_cells.ranges),
                f"celdas combinadas de {nombre}")
            self.assertEqual(
                {c: d.width for c, d in hoja_a.column_dimensions.items()},
                {c: d.width for c, d in hoja_b.column_dimensions.items()},
                f"anchos de columna de {nombre}")
            for fila in range(1, max(hoja_a.max_row, hoja_b.max_row) + 1):
                for columna in range(1, max(hoja_a.max_column, hoja_b.max_column) + 1):
                    celda_a = hoja_a.cell(row=fila, column=columna)
                    celda_b = hoja_b.cell(row=fila, column=columna)
                    donde = f"{nombre}!{celda_a.coordinate}"
                    self.assertEqual(celda_a.value, celda_b.value, donde)
                    self.assertEqual(celda_a.number_format, celda_b.number_format, donde)

    def _libro_grande(self) -> Libro:
        """Más de todo de lo que la plantilla trae de fábrica: 600 movimientos,
        12 categorías, 12 activos, 70 aportaciones gratis y 40 valoraciones."""
        libro = Libro()
        libro.ajustes.saldo_inicial = 5000.0
        libro.ajustes.objetivo_inversion = 200.0
        libro.categorias = [Categoria(nombre="Sueldo", tipo=INGRESO)]
        libro.categorias += [
            Categoria(nombre=f"Gasto {n:02d}", tipo=GASTO, presupuesto=float(10 * n))
            for n in range(1, 11)
        ]
        libro.categorias.append(Categoria(nombre="Inversión", tipo=INVERSION))
        libro.activos = [
            Activo(nombre=f"Activo {n:02d}", aportacion_inicial=float(n),
                   valor_mercado=float(100 + n), ultima_valoracion="2026-12-31")
            for n in range(1, 13)
        ]
        for i in range(600):
            categoria = libro.categorias[i % len(libro.categorias)]
            libro.movimientos.append(Movimiento(
                fecha=f"2026-{i // 50 + 1:02d}-{i % 28 + 1:02d}",
                descripcion=f"Apunte {i:03d}",
                categoria=categoria.nombre,
                importe=float(10 + i % 90),
                activo="Activo 01" if categoria.tipo == INVERSION else "",
            ))
        libro.aportaciones_gratis = [
            AportacionGratis(fecha=f"2026-{i % 12 + 1:02d}-15",
                             activo=f"Activo {i % 12 + 1:02d}",
                             concepto=f"Cashback {i:02d}", importe=1.0 + i)
            for i in range(70)
        ]
        libro.historico = [
            Valoracion(fecha=f"2026-{i % 12 + 1:02d}-{i // 12 + 1:02d}",
                       valor_mercado=1000.0 + i)
            for i in range(40)
        ]
        return libro

    @unittest.skipUnless(PLANTILLA.exists(), "no está la plantilla al lado")
    def test_las_regiones_crecen_cuando_el_libro_no_cabe(self):
        from openpyxl import load_workbook

        original = self._libro_grande()
        destino = self.carpeta / "grande.xlsx"
        _, avisos = excel.exportar(destino, original, 2026)
        self.assertEqual(avisos, [])

        cuaderno = load_workbook(destino)
        movimientos = cuaderno["Movimientos"]

        # El panel tiene las doce categorías, con la de Inversión la última.
        nombres = [movimientos.cell(row=f, column=10).value for f in range(6, 18)]
        self.assertEqual(nombres[0], "Sueldo")
        self.assertEqual(nombres[-1], "Inversión")
        self.assertEqual(len([n for n in nombres if n]), 12)

        # Las fórmulas abarcan las 600 filas nuevas y señalan a la categoría
        # de Inversión en su fila nueva, no en la de fábrica.
        self.assertIn("$E$2:$E$601", movimientos["L6"].value)
        self.assertEqual(movimientos["L19"].value,
                         '=SUMIF($K$6:$K$17,"Ingreso",L$6:L$17)')
        self.assertIn('"<>"&Movimientos!$J$17', cuaderno["Resumen"]["C4"].value)
        self.assertEqual(cuaderno["Inversiones"]["D6"].value,
                         '=IF($A6="","",SUMIFS($D$39:$D$108,$B$39:$B$108,$A6))')

        # Y en todo el libro no queda ni un límite viejo. (El $J$14 de
        # fábrica no se puede buscar a ciegas: con el panel crecido esa fila
        # es una categoría normal y aparecer, aparece.)
        rancios = ("$501", "$96")
        for nombre in cuaderno.sheetnames:
            hoja = cuaderno[nombre]
            for fila_de_celdas in hoja.iter_rows():
                for celda in fila_de_celdas:
                    valor = celda.value
                    if not isinstance(valor, str) or not valor.startswith("="):
                        continue
                    for rancio in rancios:
                        self.assertNotIn(rancio, valor,
                                         f"{nombre}!{celda.coordinate}: {valor}")

        # El presupuesto tiene una fila por categoría de gasto y su TOTAL
        # apunta al rango crecido.
        presupuesto = cuaderno["Presupuesto"]
        self.assertEqual(presupuesto["A5"].value, "=Movimientos!$J$7")
        self.assertEqual(presupuesto["A14"].value, "=Movimientos!$J$16")
        self.assertEqual(presupuesto["A15"].value, "TOTAL")
        self.assertEqual(presupuesto["B15"].value, "=SUM(B5:B14)")

        # El total de la cartera baja con los doce activos.
        inversiones = cuaderno["Inversiones"]
        self.assertEqual(inversiones["A18"].value, "TOTAL CARTERA")
        self.assertEqual(inversiones["B18"].value, "=SUM(B6:B17)")
        self.assertEqual(inversiones["K48"].value, "DE DÓNDE VIENE EL VALOR")

        # Y con todo eso, la ida y vuelta sigue siendo sin pérdidas.
        recuperado, _ = excel.importar(destino)
        self.assertEqual(retrato(recuperado), retrato(original))

    @unittest.skipUnless(PLANTILLA.exists(), "no está la plantilla al lado")
    def test_varias_categorias_de_inversion_avisan_pero_exportan(self):
        from openpyxl import load_workbook

        original = libro_completo()
        original.categorias.append(Categoria(nombre="Cripto", tipo=INVERSION))
        destino = self.carpeta / "dos-inversiones.xlsx"
        _, avisos = excel.exportar(destino, original, 2026)
        self.assertEqual(len(avisos), 1)
        self.assertIn("Cripto", avisos[0])

        # La primera hace de Inversión en la última fila del panel y las
        # demás quedan justo encima.
        cuaderno = load_workbook(destino)
        movimientos = cuaderno["Movimientos"]
        panel = [movimientos.cell(row=f, column=10).value for f in range(6, 16)]
        self.assertEqual(panel[-1], "Inversión")
        self.assertEqual(panel[-2], "Cripto")

        recuperado, _ = excel.importar(destino)
        tipos = {c.nombre: c.tipo for c in recuperado.categorias}
        self.assertEqual(tipos["Cripto"], INVERSION)
        self.assertEqual(tipos["Inversión"], INVERSION)


class PruebasAlmacen(unittest.TestCase):
    def setUp(self):
        self.carpeta = Path(tempfile.mkdtemp(prefix="contaxcell-almacen-"))
        self.almacen = Almacen(self.carpeta)

    def test_primer_arranque_crea_el_archivo(self):
        libro = self.almacen.cargar()
        self.assertTrue(self.almacen.ruta.exists())
        self.assertTrue(libro.categorias)
        self.assertEqual(self.almacen.aviso_de_arranque, "")

    def test_guarda_y_recupera(self):
        self.almacen.cargar()
        self.almacen.libro.movimientos.append(
            Movimiento(fecha="2026-06-01", categoria="Sueldo", importe=10.0))
        self.almacen.guardar()

        otro = Almacen(self.carpeta)
        otro.cargar()
        self.assertEqual(len(otro.libro.movimientos), 1)

    def test_no_deja_archivos_temporales(self):
        self.almacen.cargar()
        self.almacen.guardar()
        self.assertEqual(list(self.carpeta.glob("*.tmp")), [])

    def test_archivo_ilegible_se_aparta_y_avisa(self):
        self.almacen.ruta.write_text("{esto no es json", encoding="utf-8")
        libro = self.almacen.cargar()
        self.assertTrue(libro.categorias)
        self.assertIn("dañado", self.almacen.aviso_de_arranque)
        # Lo importante: el original no se ha perdido.
        self.assertTrue(list(self.carpeta.glob("datos-ilegible-*.json")))

    def test_copias_de_seguridad(self):
        self.almacen.cargar()
        primera = self.almacen.copia_de_seguridad("prueba")
        self.assertIsNotNone(primera)
        self.assertTrue(primera.exists())
        self.assertEqual(len(self.almacen.listar_copias()), 1)

    def test_solo_se_guardan_las_ultimas_copias(self):
        from contaxcell import almacen as modulo
        self.almacen.cargar()
        for numero in range(modulo.COPIAS_QUE_GUARDAMOS + 5):
            # El nombre lleva la hora al segundo, así que se fuerza que sean
            # distintas para que la prueba no dependa de lo rápido que vaya.
            destino = self.almacen.ruta_copias / f"2026-01-01_00-00-{numero:02d}-prueba.json"
            destino.parent.mkdir(parents=True, exist_ok=True)
            destino.write_text("{}", encoding="utf-8")
        self.almacen._limpiar_copias()
        self.assertEqual(len(self.almacen.listar_copias()), modulo.COPIAS_QUE_GUARDAMOS)

    def test_restaurar_guarda_antes_lo_que_habia(self):
        self.almacen.cargar()
        self.almacen.libro.movimientos.append(
            Movimiento(fecha="2026-06-01", categoria="Sueldo", importe=99.0))
        self.almacen.guardar()

        otro_libro = Libro.vacio()
        origen = self.carpeta / "traido.json"
        origen.write_text(json.dumps(otro_libro.a_json()), encoding="utf-8")

        self.almacen.restaurar(origen)
        self.assertEqual(self.almacen.libro.movimientos, [])
        self.assertTrue(any("antes-de-restaurar" in c.name
                            for c in self.almacen.listar_copias()))


if __name__ == "__main__":
    unittest.main()
