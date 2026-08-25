"""Puente con Excel: traer la contabilidad de la plantilla y volcarla a .xlsx.

Importar y exportar usan la misma disposición de celdas que la plantilla
original, así que un archivo exportado se puede volver a importar, y el que
descargues de Google Sheets se lee sin tocar nada.

Los lectores no van a ciegas por número de fila: buscan las cabeceras y leen
hacia abajo hasta que se acaban los datos. Así aguantan que la hoja tenga más
categorías de las previstas o alguna fila insertada por el camino.

La exportación rellena una copia de la plantilla original, con sus fórmulas
vivas, sus textos y sus colores: el archivo que sale es la hoja de siempre,
lista para seguir usándola en Excel o en Google Sheets. Si el libro trae más
datos de los que la plantilla tiene previstos, las regiones crecen y todas
las fórmulas se vuelven a generar con los límites nuevos.
"""

from __future__ import annotations

import datetime as dt
import sys
from copy import copy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection

from .modelo import (
    GASTO, INGRESO, INVERSION,
    Activo, AportacionGratis, Categoria, Libro, Movimiento, Valoracion,
    anio_de, es_fecha, redondea,
)

# Hasta dónde miramos al leer. Son topes de seguridad para no recorrer una
# hoja de un millón de filas si el archivo viene raro.
_MAX_FILAS = 20000
_MAX_BLOQUE = 500


class ErrorDeImportacion(Exception):
    """Algo que el usuario puede entender y arreglar."""


# --- conversión de celdas --------------------------------------------------

def _a_fecha(valor) -> str:
    """Cualquier cosa que parezca una fecha, en 'AAAA-MM-DD'."""
    if isinstance(valor, dt.datetime):
        return valor.date().isoformat()
    if isinstance(valor, dt.date):
        return valor.isoformat()
    if isinstance(valor, str):
        texto = valor.strip()
        if es_fecha(texto[:10]):
            return texto[:10]
        # Formato español: 24/08/2026, 24-8-26.
        for separador in ("/", "-", "."):
            partes = texto.split(separador)
            if len(partes) == 3 and all(p.strip().isdigit() for p in partes):
                dia, mes, anio = (int(p) for p in partes)
                if anio < 100:
                    anio += 2000
                try:
                    return dt.date(anio, mes, dia).isoformat()
                except ValueError:
                    return ""
    return ""


def _a_texto(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (dt.datetime, dt.date)):
        return _a_fecha(valor)
    return str(valor).strip()


def _a_numero(valor) -> float:
    if isinstance(valor, bool):
        return 0.0
    if isinstance(valor, (int, float)):
        return redondea(valor)
    texto = _a_texto(valor)
    if not texto:
        return 0.0
    limpio = "".join(c for c in texto if c.isdigit() or c in ",.-").replace(",", ".")
    try:
        return redondea(float(limpio))
    except ValueError:
        return 0.0


def _celda(hoja, fila: int, columna: int):
    return hoja.cell(row=fila, column=columna).value


def _es_total(texto: str) -> bool:
    return texto.upper().startswith("TOTAL")


def _buscar_fila(hoja, columna: int, comienza_por: str, desde: int = 1, hasta: int = 120) -> int:
    """Número de la fila cuya celda empieza por ese texto, o 0 si no está."""
    objetivo = comienza_por.upper()
    for fila in range(desde, hasta + 1):
        if _a_texto(_celda(hoja, fila, columna)).upper().startswith(objetivo):
            return fila
    return 0


# --- lectura ---------------------------------------------------------------

def importar(ruta: Path | str) -> tuple[Libro, list[str]]:
    """Devuelve el libro leído y una lista de avisos que conviene enseñar."""
    try:
        # data_only=True da el último resultado calculado de las fórmulas, que
        # es lo que nos interesa: no queremos las fórmulas, queremos los datos.
        cuaderno = load_workbook(ruta, data_only=True)
    except Exception as error:  # openpyxl lanza de todo con archivos raros
        raise ErrorDeImportacion(
            "No se ha podido abrir el archivo. Comprueba que es un .xlsx y que "
            f"no lo tienes abierto en Excel ahora mismo.\n\n({error})"
        ) from error

    try:
        if "Movimientos" not in cuaderno.sheetnames:
            raise ErrorDeImportacion(
                "El archivo no tiene ninguna hoja llamada «Movimientos».\n\n"
                "¿Seguro que es la contabilidad y no otro Excel? Si la tuya está en "
                "Google Sheets, descárgala con Archivo → Descargar → Microsoft Excel."
            )

        libro = Libro()
        avisos: list[str] = []
        movimientos = cuaderno["Movimientos"]

        libro.ajustes.saldo_inicial = _a_numero(movimientos["K3"].value)
        libro.categorias = _leer_categorias(movimientos)
        libro.movimientos = _leer_movimientos(movimientos)
        _completar_categorias_sueltas(libro, avisos)

        if "Presupuesto" in cuaderno.sheetnames:
            _leer_presupuesto(cuaderno["Presupuesto"], libro)
        if "Inversiones" in cuaderno.sheetnames:
            _leer_inversiones(cuaderno["Inversiones"], libro)
    finally:
        cuaderno.close()

    if not libro.movimientos:
        avisos.append(
            "No se ha encontrado ningún movimiento con fecha e importe. Revisa que la "
            "hoja tenga las fechas en la columna A y los importes en la E o la F."
        )
    return libro, avisos


def _leer_categorias(hoja) -> list[Categoria]:
    """Panel de control: nombres en J6 hacia abajo y su tipo en la columna K."""
    categorias = []
    for fila in range(6, 40):
        nombre = _a_texto(_celda(hoja, fila, 10))
        if not nombre or _es_total(nombre):
            break
        crudo = _a_texto(_celda(hoja, fila, 11)).lower()
        if "ingreso" in crudo:
            tipo = INGRESO
        elif "inversi" in crudo:
            tipo = INVERSION
        else:
            tipo = GASTO
        categorias.append(Categoria(nombre=nombre, tipo=tipo))
    return categorias or Libro.vacio().categorias


def _leer_movimientos(hoja) -> list[Movimiento]:
    """Fecha en A, descripción en C, categoría en D, importe en E (ingreso) o
    F (gasto), activo en H. B (Mes) y G (Balance) son fórmulas y se ignoran."""
    movimientos = []
    for fila in range(2, min(hoja.max_row, _MAX_FILAS) + 1):
        fecha = _a_fecha(_celda(hoja, fila, 1))
        if not fecha:
            continue
        ingreso = _a_numero(_celda(hoja, fila, 5))
        gasto = _a_numero(_celda(hoja, fila, 6))
        importe = abs(ingreso or gasto)
        if not importe:
            continue
        movimientos.append(Movimiento(
            fecha=fecha,
            descripcion=_a_texto(_celda(hoja, fila, 3)),
            categoria=_a_texto(_celda(hoja, fila, 4)),
            importe=importe,
            activo=_a_texto(_celda(hoja, fila, 8)),
        ))
    return movimientos


def _completar_categorias_sueltas(libro: Libro, avisos: list[str]) -> None:
    """Una categoría que aparezca en los movimientos pero no en el panel se
    añade como gasto, para que ninguna fila quede huérfana."""
    conocidas = {c.nombre for c in libro.categorias}
    for movimiento in libro.movimientos:
        nombre = movimiento.categoria
        if nombre and nombre not in conocidas:
            conocidas.add(nombre)
            libro.categorias.append(Categoria(nombre=nombre, tipo=GASTO))
            avisos.append(
                f"La categoría «{nombre}» no estaba en el panel de la hoja: la he "
                "añadido como gasto. Cámbiale el tipo en Ajustes si no lo era."
            )


def _leer_presupuesto(hoja, libro: Libro) -> None:
    """Topes mensuales por categoría y objetivo de inversión.

    Los nombres de la columna A pueden ser fórmulas que apuntan al panel de
    Movimientos. Si el archivo no trae el resultado calculado, tiramos de esa
    correspondencia: la fila n-ésima es la categoría de gasto n-ésima.
    """
    cabecera = _buscar_fila(hoja, 1, "CATEGOR") or 4
    gastos = [c for c in libro.categorias if c.tipo == GASTO]
    por_nombre = {c.nombre: c for c in libro.categorias}

    # El bloque termina en la fila «TOTAL». Una fila vacía por el medio no lo
    # corta: es una categoría sin tope, y su hueco cuenta para que la
    # correspondencia por posición siga cuadrando con las de abajo.
    vacias_seguidas = 0
    for indice, fila in enumerate(range(cabecera + 1, cabecera + 1 + _MAX_BLOQUE)):
        nombre = _a_texto(_celda(hoja, fila, 1))
        celda_importe = _celda(hoja, fila, 2)
        if _es_total(nombre):
            break
        if not nombre and celda_importe is None:
            vacias_seguidas += 1
            if vacias_seguidas >= 3:
                break
        else:
            vacias_seguidas = 0
        categoria = por_nombre.get(nombre)
        if categoria is None and indice < len(gastos):
            categoria = gastos[indice]
        if categoria is not None:
            categoria.presupuesto = _a_numero(celda_importe)

    objetivo = _buscar_fila(hoja, 1, "OBJETIVO DE INVERSI")
    if objetivo:
        libro.ajustes.objetivo_inversion = _a_numero(_celda(hoja, objetivo, 2))


def _leer_inversiones(hoja, libro: Libro) -> None:
    _leer_activos(hoja, libro)
    _leer_aportaciones_gratis(hoja, libro)
    _leer_historico(hoja, libro)


def _leer_activos(hoja, libro: Libro) -> None:
    """Nombre en A, aportación inicial en B, valor de mercado en F, fecha de
    la última valoración en I. Las columnas C, D, E, G y H se recalculan."""
    cabecera = _buscar_fila(hoja, 1, "ACTIVO") or 5
    for fila in range(cabecera + 1, cabecera + 1 + _MAX_BLOQUE):
        nombre = _a_texto(_celda(hoja, fila, 1))
        if not nombre or _es_total(nombre):
            break
        libro.activos.append(Activo(
            nombre=nombre,
            aportacion_inicial=_a_numero(_celda(hoja, fila, 2)),
            valor_mercado=_a_numero(_celda(hoja, fila, 6)),
            ultima_valoracion=_a_fecha(_celda(hoja, fila, 9)),
        ))


def _leer_aportaciones_gratis(hoja, libro: Libro) -> None:
    """El registro de cashback: su cabecera es la fila con FECHA en A y
    ACTIVO en B, que es lo que la distingue de la tabla de activos."""
    cabecera = 0
    for fila in range(1, 200):
        if (_a_texto(_celda(hoja, fila, 1)).upper() == "FECHA"
                and _a_texto(_celda(hoja, fila, 2)).upper().startswith("ACTIVO")):
            cabecera = fila
            break
    if not cabecera:
        return

    for fila in range(cabecera + 1, cabecera + 1 + _MAX_BLOQUE):
        fecha = _a_fecha(_celda(hoja, fila, 1))
        nombre = _a_texto(_celda(hoja, fila, 1))
        if _es_total(nombre):
            break
        importe = _a_numero(_celda(hoja, fila, 4))
        if not fecha or not importe:
            continue
        libro.aportaciones_gratis.append(AportacionGratis(
            fecha=fecha,
            activo=_a_texto(_celda(hoja, fila, 2)),
            concepto=_a_texto(_celda(hoja, fila, 3)),
            importe=importe,
        ))


def _leer_historico(hoja, libro: Libro) -> None:
    """Vive en las columnas K a N para no estorbar a las tablas de la
    izquierda: fecha en K y valor de mercado en M."""
    cabecera = _buscar_fila(hoja, 11, "FECHA") or 5
    for fila in range(cabecera + 1, cabecera + 1 + _MAX_BLOQUE):
        fecha = _a_fecha(_celda(hoja, fila, 11))
        if not fecha:
            continue
        libro.historico.append(Valoracion(
            fecha=fecha,
            valor_mercado=_a_numero(_celda(hoja, fila, 13)),
        ))


# --- escritura -------------------------------------------------------------
#
# Exportar no fabrica una hoja de cero: abre la plantilla original que viaja
# con la aplicación, la rellena con los datos del libro y la guarda. Así el
# archivo exportado ES la plantilla de siempre, con sus fórmulas vivas.
#
# La plantilla tiene sitio para un número fijo de filas en cada tabla. Si el
# libro cabe, solo se tocan las celdas de datos. Si no cabe, la región crece
# y todas las fórmulas del libro entero se vuelven a generar desde la tabla
# FORMULAS con los límites nuevos. Nunca se insertan filas a lo bruto: en
# Movimientos y en Inversiones hay regiones que comparten filas, y un
# insert_rows arrastraría a la vecina.

# Cuántos huecos trae la plantilla de fábrica.
_HUECOS_MOVIMIENTOS = 500
_HUECOS_CATEGORIAS = 9
_HUECOS_ACTIVOS = 10
_HUECOS_GRATIS = 60
_HUECOS_HISTORICO = 36

# Todas las fórmulas de la plantilla, con huecos para los límites que
# dependen del tamaño de los datos. Con los tamaños de fábrica, cada fórmula
# generada es idéntica letra a letra a la que trae la plantilla.
#   f      → la fila de la propia fórmula
#   m      → última fila de datos de Movimientos (501 de fábrica)
#   cats   → última fila del panel de categorías (14)
#   inv    → fila de la categoría de Inversión dentro del panel (14)
#   r      → fila del panel a la que apunta una referencia cruzada
#   g1, g2 → primera y última fila del registro de aportaciones gratis (37, 96)
#   t      → fila del TOTAL CARTERA de Inversiones (16)
FORMULAS = {
    # Movimientos: las dos columnas calculadas del libro diario.
    "mes_de_fila": '=IF($A{f}="","",TEXT($A{f},"mmmm yyyy"))',
    "balance_de_fila": '=IF($A{f}="","",$K$3+SUM($E$2:$E{f})-SUM($F$2:$F{f}))',

    # El panel de control: cada categoría y los totales de debajo.
    "panel_mes_activo": (
        '=IF($J{f}="","",IF($K{f}="Ingreso",'
        'SUMIFS($E$2:$E${m},$A$2:$A${m},">="&$K$2,$A$2:$A${m},"<"&EOMONTH($K$2,0)+1,$D$2:$D${m},$J{f}),'
        'SUMIFS($F$2:$F${m},$A$2:$A${m},">="&$K$2,$A$2:$A${m},"<"&EOMONTH($K$2,0)+1,$D$2:$D${m},$J{f})))'
    ),
    "panel_total_anio": (
        '=IF($J{f}="","",IF($K{f}="Ingreso",'
        'SUMIFS($E$2:$E${m},$A$2:$A${m},">="&DATE(YEAR($K$2),1,1),$A$2:$A${m},"<"&DATE(YEAR($K$2)+1,1,1),$D$2:$D${m},$J{f}),'
        'SUMIFS($F$2:$F${m},$A$2:$A${m},">="&DATE(YEAR($K$2),1,1),$A$2:$A${m},"<"&DATE(YEAR($K$2)+1,1,1),$D$2:$D${m},$J{f})))'
    ),
    "panel_media_mensual": '=IFERROR($M{f}/$L${meses},0)',
    "panel_total_tipo": '=SUMIF($K$6:$K${cats},"{tipo}",{col}$6:{col}${cats})',
    "panel_ahorro": '={col}{ingresos}-{col}{gastos}',
    "panel_tasa": '=IFERROR({col}{ahorro}/{col}{ingresos},0)',
    "panel_flujo": '={col}{ahorro}-{col}{inversion}',
    "panel_meses_con_datos":
        '=SUMPRODUCT(--((Resumen!$B$4:$B$15+Resumen!$C$4:$C$15+Resumen!$E$4:$E$15)>0))',
    "panel_saldo_banco": '=$K$3+SUM($E$2:$E${m})-SUM($F$2:$F${m})',

    # Resumen: la tabla de los doce meses (n es el número de mes, n1 el
    # siguiente), los repartos por categoría y los indicadores.
    "res_anio": "=YEAR(Movimientos!$K$2)",
    "res_mes": "=DATE($I$1,{n},1)",
    "res_ingresos": (
        '=SUMIFS(Movimientos!$E$2:$E${m},Movimientos!$A$2:$A${m},">="&DATE($I$1,{n},1),'
        'Movimientos!$A$2:$A${m},"<"&DATE($I$1,{n1},1))'
    ),
    "res_gastos": (
        '=SUMIFS(Movimientos!$F$2:$F${m},Movimientos!$A$2:$A${m},">="&DATE($I$1,{n},1),'
        'Movimientos!$A$2:$A${m},"<"&DATE($I$1,{n1},1),Movimientos!$D$2:$D${m},"<>"&Movimientos!$J${inv})'
    ),
    "res_ahorro": "=B{f}-C{f}",
    "res_inversion": (
        '=SUMIFS(Movimientos!$F$2:$F${m},Movimientos!$A$2:$A${m},">="&DATE($I$1,{n},1),'
        'Movimientos!$A$2:$A${m},"<"&DATE($I$1,{n1},1),Movimientos!$D$2:$D${m},Movimientos!$J${inv})'
    ),
    "res_saldo": (
        '=Movimientos!$K$3+SUMIFS(Movimientos!$E$2:$E${m},Movimientos!$A$2:$A${m},"<"&DATE($I$1,{n1},1))'
        '-SUMIFS(Movimientos!$F$2:$F${m},Movimientos!$A$2:$A${m},"<"&DATE($I$1,{n1},1))'
    ),
    "res_etiqueta": '=TEXT(A{f},"mmm")',
    "res_total_columna": "=SUM({col}4:{col}15)",
    "res_saldo_final": "=Movimientos!$L${saldo}",
    "res_cat_nombre": "=Movimientos!$J${r}",
    "res_cat_importe": "=Movimientos!$M${r}",
    "res_cat_pct": "=IFERROR(I{f}/$I${total},0)",
    "res_bloque_total": "=SUM(I{desde}:I{hasta})",
    "res_bloque_pct": "=IFERROR(SUM(J{desde}:J{hasta}),0)",
    "res_cuadre": "=I{total}-C16",
    "res_ahorro_medio": "=IFERROR($D$16/Movimientos!$L${meses},0)",
    "res_gasto_medio": "=IFERROR($C$16/Movimientos!$L${meses},0)",
    "res_tasa_anual": "=IFERROR($D$16/$B$16,0)",
    "res_mayor_gasto":
        '=IFERROR(TEXT(INDEX($A$4:$A$15,MATCH(MAX($C$4:$C$15),$C$4:$C$15,0)),"mmmm"),"")',
    "res_colchon": "=IFERROR(Movimientos!$L${saldo}/($C$16/Movimientos!$L${meses}),0)",
    "res_inversion_anual": "=$E$16",
    "res_cartera": "=Inversiones!${col}${t}",
    "res_patrimonio": "=Movimientos!$L${saldo}+Inversiones!$F${t}",

    # Presupuesto: una fila por categoría de gasto y el bloque de abajo.
    "pre_nombre": "=Movimientos!$J${r}",
    "pre_real": "=Movimientos!$L${r}",
    "pre_disponible": "=B{f}-C{f}",
    "pre_pct": "=IFERROR(C{f}/B{f},0)",
    "pre_total": "=SUM({col}{desde}:{col}{hasta})",
    "pre_ingresos": "=Movimientos!$L${ingresos}",
    "pre_margen": "=B{ingresos}-B{total}",
    "pre_aportado": "=Movimientos!$L${inversion}",
    "pre_pendiente": "=MAX(0,B{objetivo}-B{aportado})",

    # Inversiones: la tabla de activos, sus totales y las notas laterales.
    "act_banco": (
        '=IF($A{f}="","",SUMIFS(Movimientos!$F$2:$F${m},'
        'Movimientos!$D$2:$D${m},Movimientos!$J${inv},Movimientos!$H$2:$H${m},$A{f}))'
    ),
    "act_gratis": '=IF($A{f}="","",SUMIFS($D${g1}:$D${g2},$B${g1}:$B${g2},$A{f}))',
    "act_aportado": '=IF($A{f}="","",$B{f}+$C{f}+$D{f})',
    "act_generado": '=IF($A{f}="","",$F{f}-$E{f})',
    "act_rentabilidad": '=IF(OR($A{f}="",$E{f}=0),"",$G{f}/$E{f})',
    "cartera_suma": "=SUM({col}6:{col}{hasta})",
    "cartera_generado": "=F{t}-E{t}",
    "cartera_rentabilidad": "=IFERROR(G{t}/E{t},0)",
    "banco_sin_asignar": (
        "=SUMIFS(Movimientos!$F$2:$F${m},Movimientos!$D$2:$D${m},Movimientos!$J${inv})-C{t}"
    ),
    "gratis_sin_asignar": "=SUM($D${g1}:$D${g2})-D{t}",
    "cartera_indicador": "={col}{t}",
    "ganado_sin_poner": "=G{t}+D{t}",
    "aportado_banco_anio": '=SUMIF(Movimientos!$K$6:$K${cats},"Inversión",Movimientos!$M$6:$M${cats})',
    "aportado_banco_mes": '=SUMIF(Movimientos!$K$6:$K${cats},"Inversión",Movimientos!$L$6:$L${cats})',
    "regla_gasto_mes": "=Movimientos!$L${gastos}",
    "regla_estimacion": "=$G${gasto}*$G${pct}",
    "gratis_total": "=SUM(D{g1}:D{g2})",

    # El histórico de la cartera, en las columnas K a N de Inversiones.
    "hist_aportado": (
        '=IF($K{f}="","",$B${t}+SUMIFS(Movimientos!$F$2:$F${m},'
        'Movimientos!$D$2:$D${m},Movimientos!$J${inv},Movimientos!$A$2:$A${m},"<="&$K{f})'
        '+SUMIFS($D${g1}:$D${g2},$A${g1}:$A${g2},"<="&$K{f}))'
    ),
    "hist_generado": '=IF($K{f}="","",$M{f}-$L{f})',
}


def _formula(nombre: str, **huecos) -> str:
    return FORMULAS[nombre].format(**huecos)


def _ruta_plantilla() -> Path:
    """La plantilla viaja con el programa, igual que el icono.

    Empaquetado con PyInstaller, los recursos se descomprimen en la carpeta
    que el propio ejecutable anuncia en `sys._MEIPASS`.
    """
    if getattr(sys, "frozen", False):
        return Path(getattr(sys, "_MEIPASS", ".")) / "recursos" / "plantilla.xlsx"
    return Path(__file__).resolve().parent.parent / "recursos" / "plantilla.xlsx"


@dataclass
class _Disposicion:
    """Dónde empieza y acaba cada región una vez decidido cuánto crece."""

    fin_movimientos: int    # última fila de datos de Movimientos (501)
    fin_categorias: int     # última fila física del panel de categorías (14)
    fila_inversion: int     # la fila de la categoría de Inversión (14)
    salto_panel: int        # cuántas filas baja lo que hay bajo el panel (0)
    fila_cartera: int       # la fila del TOTAL CARTERA de Inversiones (16)
    salto_activos: int      # cuántas filas baja lo que hay bajo los activos (0)
    inicio_gratis: int      # primera fila del registro de cashback (37)
    fin_gratis: int         # última fila del registro de cashback (96)
    fin_historico: int      # última fila de datos del histórico (41)
    salto_historico: int    # cuánto baja el bloque de debajo del histórico (0)


def _disposicion(libro: Libro, categorias: list[Categoria]) -> _Disposicion:
    hay_inversion = any(c.tipo == INVERSION for c in categorias)
    # Sin categoría de inversión, las referencias del libro apuntan a la
    # primera fila libre del panel, que se queda en blanco a propósito.
    huecos_cat = max(_HUECOS_CATEGORIAS, len(categorias) + (0 if hay_inversion else 1))
    huecos_act = max(_HUECOS_ACTIVOS, len(libro.activos))
    huecos_gratis = max(_HUECOS_GRATIS, len(libro.aportaciones_gratis))
    huecos_hist = max(_HUECOS_HISTORICO, len(libro.historico))
    salto_activos = huecos_act - _HUECOS_ACTIVOS
    return _Disposicion(
        fin_movimientos=1 + max(_HUECOS_MOVIMIENTOS, len(libro.movimientos)),
        fin_categorias=5 + huecos_cat,
        fila_inversion=5 + len(categorias) + (0 if hay_inversion else 1),
        salto_panel=huecos_cat - _HUECOS_CATEGORIAS,
        fila_cartera=6 + huecos_act,
        salto_activos=salto_activos,
        inicio_gratis=37 + salto_activos,
        fin_gratis=36 + salto_activos + huecos_gratis,
        fin_historico=5 + huecos_hist,
        salto_historico=huecos_hist - _HUECOS_HISTORICO,
    )


def _categorias_del_panel(libro: Libro) -> tuple[list[Categoria], list[str]]:
    """El orden del panel: la categoría de Inversión va la última, porque
    todo el libro la señala por su celda. Si hay más de una de ese tipo, la
    plantilla solo puede tratar como inversión a la primera."""
    inversiones = [c for c in libro.categorias if c.tipo == INVERSION]
    normales = [c for c in libro.categorias if c.tipo != INVERSION]
    avisos: list[str] = []
    if len(inversiones) > 1:
        nombres = ", ".join(f"«{c.nombre}»" for c in inversiones[1:])
        avisos.append(
            f"La plantilla solo contempla una categoría de tipo Inversión y el libro "
            f"tiene {len(inversiones)}. En el Excel hará de inversión «{inversiones[0].nombre}»; "
            f"{nombres} se apuntan en el panel, pero sus fórmulas no las descuentan como inversión."
        )
    return normales + inversiones[1:] + inversiones[:1], avisos


def _mes_activo(libro: Libro, anio: int) -> dt.date:
    """El día 1 del mes del último movimiento de ese año. Sin movimientos en
    el año, diciembre, que es lo que deja el año entero a la vista."""
    fechas = [m.fecha for m in libro.movimientos if anio_de(m.fecha) == anio]
    if fechas:
        return dt.date(anio, int(max(fechas)[5:7]), 1)
    return dt.date(anio, 12, 1)


def _fecha_excel(iso: str):
    return dt.date.fromisoformat(iso) if es_fecha(iso) else None


# Copiar y limpiar celdas conservando el aspecto. openpyxl comparte los
# objetos de estilo entre celdas, así que siempre se trabaja con copias.

def _foto_estilo(celda) -> tuple:
    return (copy(celda.font), copy(celda.fill), copy(celda.border),
            copy(celda.alignment), celda.number_format, copy(celda.protection))

def _pega_estilo(celda, estilo: tuple) -> None:
    celda.font, celda.fill, celda.border, celda.alignment = (
        copy(estilo[0]), copy(estilo[1]), copy(estilo[2]), copy(estilo[3]))
    celda.number_format = estilo[4]
    celda.protection = copy(estilo[5])

def _foto_bloque(hoja, fila1: int, fila2: int, col1: int, col2: int) -> list:
    """Valores y estilos de un rectángulo, para volver a pegarlo en otro sitio."""
    return [
        [(hoja.cell(row=f, column=c).value, _foto_estilo(hoja.cell(row=f, column=c)))
         for c in range(col1, col2 + 1)]
        for f in range(fila1, fila2 + 1)
    ]

def _pega_bloque(hoja, fila: int, col1: int, foto: list) -> None:
    for df, fila_foto in enumerate(foto):
        for dc, (valor, estilo) in enumerate(fila_foto):
            celda = hoja.cell(row=fila + df, column=col1 + dc)
            celda.value = valor
            _pega_estilo(celda, estilo)

def _limpia_bloque(hoja, fila1: int, fila2: int, col1: int, col2: int) -> None:
    for f in range(fila1, fila2 + 1):
        for c in range(col1, col2 + 1):
            celda = hoja.cell(row=f, column=c)
            celda.value = None
            celda.font = Font()
            celda.fill = PatternFill()
            celda.border = Border()
            celda.alignment = Alignment()
            celda.number_format = "General"
            celda.protection = Protection()


def _pon(hoja, fila: int, columna: int, valor) -> None:
    hoja.cell(row=fila, column=columna).value = valor


def exportar(ruta: Path | str, libro: Libro, anio: int) -> tuple[Path, list[str]]:
    """Rellena una copia de la plantilla con el libro y la guarda en `ruta`.

    Devuelve la ruta escrita y una lista de avisos (normalmente vacía)."""
    categorias, avisos = _categorias_del_panel(libro)
    d = _disposicion(libro, categorias)
    cuaderno = load_workbook(_ruta_plantilla())
    try:
        _volcar_movimientos(cuaderno["Movimientos"], libro, categorias, d, anio)
        _volcar_resumen(cuaderno["Resumen"], categorias, d)
        _volcar_presupuesto(cuaderno["Presupuesto"], libro, categorias, d)
        _volcar_inversiones(cuaderno["Inversiones"], libro, d)
        destino = Path(ruta)
        cuaderno.save(destino)
    finally:
        cuaderno.close()
    return destino, avisos


def _volcar_movimientos(hoja, libro: Libro, categorias: list[Categoria],
                        d: _Disposicion, anio: int) -> None:
    m = d.fin_movimientos
    salto = d.salto_panel

    # Los estilos de referencia se fotografían antes de tocar nada: filas de
    # datos pares e impares (van a rayas) y una fila del panel por cada tipo.
    proto_datos = {
        paridad: [_foto_estilo(hoja.cell(row=fila, column=c)) for c in range(1, 9)]
        for paridad, fila in ((0, 500), (1, 501))
    }
    proto_panel = {
        tipo: [_foto_estilo(hoja.cell(row=fila, column=c)) for c in range(10, 15)]
        for tipo, fila in ((INGRESO, 6), (GASTO, 8), (INVERSION, 14))
    }
    pie = _foto_bloque(hoja, 15, 36, 10, 14)  # totales y «CÓMO USARLO»

    # Las dos celdas amarillas del panel.
    _pon(hoja, 2, 11, _mes_activo(libro, anio))
    _pon(hoja, 3, 11, libro.ajustes.saldo_inicial)

    # El libro diario. El orden por fecha es estable: dos apuntes del mismo
    # día conservan el orden en que estaban.
    movimientos = sorted(libro.movimientos, key=lambda mov: mov.fecha)
    for indice, fila in enumerate(range(2, m + 1)):
        if fila > 501:
            for c in range(1, 9):
                _pega_estilo(hoja.cell(row=fila, column=c), proto_datos[fila % 2][c - 1])
        mov = movimientos[indice] if indice < len(movimientos) else None
        es_ingreso = mov is not None and libro.tipo_de(mov.categoria) == INGRESO
        _pon(hoja, fila, 1, _fecha_excel(mov.fecha) if mov else None)
        _pon(hoja, fila, 2, _formula("mes_de_fila", f=fila))
        _pon(hoja, fila, 3, (mov.descripcion or None) if mov else None)
        _pon(hoja, fila, 4, (mov.categoria or None) if mov else None)
        _pon(hoja, fila, 5, mov.importe if mov and es_ingreso else None)
        _pon(hoja, fila, 6, mov.importe if mov and not es_ingreso else None)
        _pon(hoja, fila, 7, _formula("balance_de_fila", f=fila))
        _pon(hoja, fila, 8, (mov.activo or None) if mov else None)

    # El panel de categorías. Primero se despeja todo lo que hay debajo de
    # las cabeceras (descombinando las celdas del texto, que si no son de
    # solo lectura), luego se escriben las categorías y se vuelve a pegar el
    # pie donde le toque.
    for fila in range(27, 37):
        hoja.unmerge_cells(f"J{fila}:N{fila}")
    _limpia_bloque(hoja, 15, 36 + max(salto, 0), 10, 14)
    for indice, fila in enumerate(range(6, d.fin_categorias + 1)):
        cat = categorias[indice] if indice < len(categorias) else None
        if cat is not None:
            for c in range(10, 15):
                _pega_estilo(hoja.cell(row=fila, column=c), proto_panel[cat.tipo][c - 10])
        _pon(hoja, fila, 10, cat.nombre if cat else None)
        _pon(hoja, fila, 11, cat.tipo if cat else None)
        _pon(hoja, fila, 12, _formula("panel_mes_activo", f=fila, m=m))
        _pon(hoja, fila, 13, _formula("panel_total_anio", f=fila, m=m))
        _pon(hoja, fila, 14, _formula("panel_media_mensual", f=fila, meses=23 + salto))
    _pega_bloque(hoja, 15 + salto, 10, pie)
    for fila in range(27, 37):
        hoja.merge_cells(f"J{fila + salto}:N{fila + salto}")

    # Las fórmulas del pie, con los límites que tocan.
    cats = d.fin_categorias
    ingresos, gastos, ahorro, inversion = (16 + salto, 17 + salto, 18 + salto, 20 + salto)
    for col in ("L", "M"):
        hoja[f"{col}{ingresos}"] = _formula("panel_total_tipo", cats=cats, tipo="Ingreso", col=col)
        hoja[f"{col}{gastos}"] = _formula("panel_total_tipo", cats=cats, tipo="Gasto", col=col)
        hoja[f"{col}{ahorro}"] = _formula("panel_ahorro", col=col, ingresos=ingresos, gastos=gastos)
        hoja[f"{col}{ahorro + 1}"] = _formula("panel_tasa", col=col, ahorro=ahorro, ingresos=ingresos)
        hoja[f"{col}{inversion}"] = _formula("panel_total_tipo", cats=cats, tipo="Inversión", col=col)
        hoja[f"{col}{inversion + 1}"] = _formula("panel_flujo", col=col, ahorro=ahorro, inversion=inversion)
    hoja[f"L{23 + salto}"] = _formula("panel_meses_con_datos")
    hoja[f"L{24 + salto}"] = _formula("panel_saldo_banco", m=m)

    hoja.auto_filter.ref = f"A1:H{m}"
    for dv in hoja.data_validations.dataValidation:
        formula = dv.formula1 or ""
        if dv.type == "decimal":
            dv.sqref = f"E2:F{m}"
        elif "Inversiones" in formula:
            dv.formula1 = f"Inversiones!$A$6:$A${d.fila_cartera - 1}"
            dv.sqref = f"H2:H{m}"
        else:
            dv.formula1 = f"Movimientos!$J$6:$J${d.fila_inversion}"
            dv.sqref = f"D2:D{m}"


def _volcar_resumen(hoja, categorias: list[Categoria], d: _Disposicion) -> None:
    m = d.fin_movimientos
    inv = d.fila_inversion
    saldo, meses = 24 + d.salto_panel, 23 + d.salto_panel

    hoja["I1"] = _formula("res_anio")

    # Los doce meses. Sus filas no se mueven nunca: el panel de Movimientos
    # cuenta con encontrarlas en B4:B15.
    for numero, fila in enumerate(range(4, 16), start=1):
        hoja[f"A{fila}"] = _formula("res_mes", n=numero)
        hoja[f"B{fila}"] = _formula("res_ingresos", m=m, n=numero, n1=numero + 1)
        hoja[f"C{fila}"] = _formula("res_gastos", m=m, n=numero, n1=numero + 1, inv=inv)
        hoja[f"D{fila}"] = _formula("res_ahorro", f=fila)
        hoja[f"E{fila}"] = _formula("res_inversion", m=m, n=numero, n1=numero + 1, inv=inv)
        hoja[f"F{fila}"] = _formula("res_saldo", m=m, n1=numero + 1)
        hoja[f"M{fila}"] = _formula("res_etiqueta", f=fila)
    for col in ("B", "C", "D", "E"):
        hoja[f"{col}16"] = _formula("res_total_columna", col=col)
    hoja["F16"] = _formula("res_saldo_final", saldo=saldo)

    # Los repartos por categoría de las columnas H a J. Crecen con las
    # categorías, así que se reconstruyen enteros.
    filas_gasto = [6 + i for i, c in enumerate(categorias) if c.tipo == GASTO]
    filas_ingreso = [6 + i for i, c in enumerate(categorias) if c.tipo == INGRESO]
    n_gasto, n_ingreso = max(1, len(filas_gasto)), max(1, len(filas_ingreso))
    total_gasto = 4 + n_gasto
    cabecera_ingreso = total_gasto + 3
    total_ingreso = cabecera_ingreso + n_ingreso + 1

    proto_gasto = [_foto_estilo(hoja.cell(row=4, column=c)) for c in range(8, 11)]
    proto_ingreso = [_foto_estilo(hoja.cell(row=14, column=c)) for c in range(8, 11)]
    foto_total_gasto = _foto_bloque(hoja, 10, 10, 8, 10)
    foto_cuadre = _foto_bloque(hoja, 11, 11, 8, 10)
    foto_cabecera_ingreso = _foto_bloque(hoja, 13, 13, 8, 10)
    foto_total_ingreso = _foto_bloque(hoja, 16, 16, 8, 10)

    _limpia_bloque(hoja, 4, max(16, total_ingreso), 8, 10)
    for bloque, filas_panel, proto, desde, total in (
        ("gasto", filas_gasto, proto_gasto, 4, total_gasto),
        ("ingreso", filas_ingreso, proto_ingreso, cabecera_ingreso + 1, total_ingreso),
    ):
        n = max(1, len(filas_panel))
        for indice, fila in enumerate(range(desde, desde + n)):
            for c in range(8, 11):
                _pega_estilo(hoja.cell(row=fila, column=c), proto[c - 8])
            if indice < len(filas_panel):
                r = filas_panel[indice]
                hoja[f"H{fila}"] = _formula("res_cat_nombre", r=r)
                hoja[f"I{fila}"] = _formula("res_cat_importe", r=r)
                hoja[f"J{fila}"] = _formula("res_cat_pct", f=fila, total=total)
        if bloque == "gasto":
            _pega_bloque(hoja, total_gasto, 8, foto_total_gasto)
            _pega_bloque(hoja, total_gasto + 1, 8, foto_cuadre)
            _pega_bloque(hoja, cabecera_ingreso, 8, foto_cabecera_ingreso)
        else:
            _pega_bloque(hoja, total_ingreso, 8, foto_total_ingreso)
        hoja[f"I{total}"] = _formula("res_bloque_total", desde=desde, hasta=desde + n - 1)
        hoja[f"J{total}"] = _formula("res_bloque_pct", desde=desde, hasta=desde + n - 1)
    hoja[f"I{total_gasto + 1}"] = _formula("res_cuadre", total=total_gasto)

    # Los indicadores de abajo a la izquierda.
    hoja["B19"] = _formula("res_saldo_final", saldo=saldo)
    hoja["B20"] = _formula("res_ahorro_medio", meses=meses)
    hoja["B21"] = _formula("res_gasto_medio", meses=meses)
    hoja["B22"] = _formula("res_tasa_anual")
    hoja["B23"] = _formula("res_mayor_gasto")
    hoja["B24"] = _formula("res_colchon", saldo=saldo, meses=meses)
    hoja["B25"] = _formula("res_inversion_anual")
    for col, fila in (("E", 26), ("F", 27), ("G", 28)):
        hoja[f"B{fila}"] = _formula("res_cartera", col=col, t=d.fila_cartera)
    hoja["B29"] = _formula("res_patrimonio", saldo=saldo, t=d.fila_cartera)


def _volcar_presupuesto(hoja, libro: Libro, categorias: list[Categoria],
                        d: _Disposicion) -> None:
    gastos = [(6 + i, c) for i, c in enumerate(categorias) if c.tipo == GASTO]
    n = max(1, len(gastos))
    total = 5 + n

    proto = [_foto_estilo(hoja.cell(row=5, column=c)) for c in range(1, 6)]
    pie = _foto_bloque(hoja, 11, 17, 1, 5)  # TOTAL y el bloque de abajo

    _limpia_bloque(hoja, 5, max(17, total + 6), 1, 5)
    for indice, fila in enumerate(range(5, 5 + n)):
        for c in range(1, 6):
            _pega_estilo(hoja.cell(row=fila, column=c), proto[c - 1])
        if indice < len(gastos):
            r, cat = gastos[indice]
            hoja[f"A{fila}"] = _formula("pre_nombre", r=r)
            hoja[f"B{fila}"] = cat.presupuesto
            hoja[f"C{fila}"] = _formula("pre_real", r=r)
            hoja[f"D{fila}"] = _formula("pre_disponible", f=fila)
            hoja[f"E{fila}"] = _formula("pre_pct", f=fila)
    _pega_bloque(hoja, total, 1, pie)

    for col in ("B", "C", "D"):
        hoja[f"{col}{total}"] = _formula("pre_total", col=col, desde=5, hasta=total - 1)
    hoja[f"E{total}"] = _formula("pre_pct", f=total)
    ingresos, margen, objetivo, aportado, pendiente = range(total + 2, total + 7)
    hoja[f"B{ingresos}"] = _formula("pre_ingresos", ingresos=16 + d.salto_panel)
    hoja[f"B{margen}"] = _formula("pre_margen", ingresos=ingresos, total=total)
    hoja[f"B{objetivo}"] = libro.ajustes.objetivo_inversion
    hoja[f"B{aportado}"] = _formula("pre_aportado", inversion=20 + d.salto_panel)
    hoja[f"B{pendiente}"] = _formula("pre_pendiente", objetivo=objetivo, aportado=aportado)


def _volcar_inversiones(hoja, libro: Libro, d: _Disposicion) -> None:
    m, inv, cats = d.fin_movimientos, d.fila_inversion, d.fin_categorias
    t, salto = d.fila_cartera, d.salto_activos
    g1, g2 = d.inicio_gratis, d.fin_gratis

    proto_activo = [_foto_estilo(hoja.cell(row=7, column=c)) for c in range(1, 10)]
    proto_gratis = [_foto_estilo(hoja.cell(row=38, column=c)) for c in range(1, 5)]
    proto_hist = [_foto_estilo(hoja.cell(row=7, column=c)) for c in range(11, 15)]
    # Del TOTAL CARTERA a la cabecera del registro de cashback, tal cual.
    cuerpo = _foto_bloque(hoja, 16, 36, 1, 9)
    regla = _foto_bloque(hoja, 37, 41, 6, 7)  # la nota del % de cashback
    fila_total_gratis = _foto_bloque(hoja, 97, 97, 1, 9)
    bloque_valor = _foto_bloque(hoja, 44, 48, 11, 12)  # «DE DÓNDE VIENE EL VALOR»

    # Las celdas combinadas de esta zona se sueltan antes de limpiar y se
    # vuelven a combinar, desplazadas, al final.
    _COMBINADAS_CARTERA = ("A34:D34", "A35:D35", "F36:G36", "F40:G41")
    for rango in _COMBINADAS_CARTERA:
        hoja.unmerge_cells(rango)

    # La tabla de activos.
    _limpia_bloque(hoja, 16, max(97, g2 + 1), 1, 9)
    for indice, fila in enumerate(range(6, t)):
        for c in range(1, 10):
            _pega_estilo(hoja.cell(row=fila, column=c), proto_activo[c - 1])
        activo = libro.activos[indice] if indice < len(libro.activos) else None
        _pon(hoja, fila, 1, activo.nombre if activo else None)
        _pon(hoja, fila, 2, activo.aportacion_inicial if activo else None)
        _pon(hoja, fila, 3, _formula("act_banco", f=fila, m=m, inv=inv))
        _pon(hoja, fila, 4, _formula("act_gratis", f=fila, g1=g1, g2=g2))
        _pon(hoja, fila, 5, _formula("act_aportado", f=fila))
        _pon(hoja, fila, 6, activo.valor_mercado if activo else None)
        _pon(hoja, fila, 7, _formula("act_generado", f=fila))
        _pon(hoja, fila, 8, _formula("act_rentabilidad", f=fila))
        _pon(hoja, fila, 9, _fecha_excel(activo.ultima_valoracion) if activo else None)
    _pega_bloque(hoja, t, 1, cuerpo)
    _pega_bloque(hoja, 37 + salto, 6, regla)

    # Las fórmulas del cuerpo pegado, con los límites que tocan.
    for col in ("B", "C", "D", "E", "F"):
        hoja[f"{col}{t}"] = _formula("cartera_suma", col=col, hasta=t - 1)
    hoja[f"G{t}"] = _formula("cartera_generado", t=t)
    hoja[f"H{t}"] = _formula("cartera_rentabilidad", t=t)
    hoja[f"B{18 + salto}"] = _formula("banco_sin_asignar", m=m, inv=inv, t=t)
    hoja[f"B{19 + salto}"] = _formula("gratis_sin_asignar", g1=g1, g2=g2, t=t)
    for col, fila in (("B", 22), ("C", 23), ("D", 24), ("E", 25), ("F", 26),
                      ("G", 27), ("H", 28)):
        hoja[f"B{fila + salto}"] = _formula("cartera_indicador", col=col, t=t)
    hoja[f"B{29 + salto}"] = _formula("ganado_sin_poner", t=t)
    hoja[f"B{30 + salto}"] = _formula("aportado_banco_anio", cats=cats)
    hoja[f"B{31 + salto}"] = _formula("aportado_banco_mes", cats=cats)
    hoja[f"G{38 + salto}"] = _formula("regla_gasto_mes", gastos=17 + d.salto_panel)
    hoja[f"G{39 + salto}"] = _formula("regla_estimacion", gasto=38 + salto, pct=37 + salto)

    # El registro de aportaciones gratis y su total.
    aportaciones = sorted(libro.aportaciones_gratis, key=lambda a: a.fecha)
    for indice, fila in enumerate(range(g1, g2 + 1)):
        for c in range(1, 5):
            _pega_estilo(hoja.cell(row=fila, column=c), proto_gratis[c - 1])
        gratis = aportaciones[indice] if indice < len(aportaciones) else None
        _pon(hoja, fila, 1, _fecha_excel(gratis.fecha) if gratis else None)
        _pon(hoja, fila, 2, (gratis.activo or None) if gratis else None)
        _pon(hoja, fila, 3, (gratis.concepto or None) if gratis else None)
        _pon(hoja, fila, 4, gratis.importe if gratis else None)
    _pega_bloque(hoja, g2 + 1, 1, fila_total_gratis)
    hoja[f"D{g2 + 1}"] = _formula("gratis_total", g1=g1, g2=g2)

    for rango in _COMBINADAS_CARTERA:
        inicio, fin = rango.split(":")
        hoja.merge_cells(f"{inicio[0]}{int(inicio[1:]) + salto}:{fin[0]}{int(fin[1:]) + salto}")

    # El histórico de la cartera, en sus columnas K a N.
    valoraciones = sorted(libro.historico, key=lambda v: v.fecha)
    _limpia_bloque(hoja, 6, max(48, 48 + d.salto_historico), 11, 14)
    for indice, fila in enumerate(range(6, d.fin_historico + 1)):
        for c in range(11, 15):
            _pega_estilo(hoja.cell(row=fila, column=c), proto_hist[c - 11])
        punto = valoraciones[indice] if indice < len(valoraciones) else None
        _pon(hoja, fila, 11, _fecha_excel(punto.fecha) if punto else None)
        _pon(hoja, fila, 12, _formula("hist_aportado", f=fila, m=m, inv=inv, t=t, g1=g1, g2=g2))
        _pon(hoja, fila, 13, punto.valor_mercado if punto else None)
        _pon(hoja, fila, 14, _formula("hist_generado", f=fila))
    _pega_bloque(hoja, 44 + d.salto_historico, 11, bloque_valor)
    for col, fila in (("B", 45), ("C", 46), ("D", 47), ("G", 48)):
        hoja[f"L{fila + d.salto_historico}"] = _formula("cartera_indicador", col=col, t=t)

    for dv in hoja.data_validations.dataValidation:
        dv.formula1 = f"$A$6:$A${t - 1}"
        dv.sqref = f"B{g1}:B{g2}"
